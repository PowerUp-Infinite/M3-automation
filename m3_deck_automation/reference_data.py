"""
reference_data.py — loads AUM, Power Rank, Upside/Downside, and Rolling Returns.
Returns (lookup, rr_category) for use when populating scheme slides.
"""
import os
import re

import openpyxl
import pandas as pd


def _norm_name(name):
    """Normalize fund name for fuzzy matching: lowercase, strip 'Fund'/'Fd', collapse spaces."""
    s = str(name).strip().lower()
    s = re.sub(r'\bfund\b', '', s)
    s = re.sub(r'\bfd\b', '', s)      # AUM file abbreviates 'Fund' as 'Fd' in some names
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def load_reference_data(project_dir):
    """
    Load all reference files. Returns (lookup, rr_category):
      lookup: dict keyed by ISIN:
        'aum_cr':    float or None  — AUM in Crores
        'powerrank': int   or None  — Power rank integer
        'upside':    float or None  — Upside capture ratio
        'downside':  float or None  — Downside capture ratio
        'five_year': float or None  — 5Y rolling return (e.g. 29.76 means 29.76%)
      rr_category: dict keyed by UPDATED_SUBCATEGORY name -> 5Y category avg return float
    """
    lookup = {}

    # --- AUM + fund-name fallback ---
    aum_path = os.path.join(project_dir, 'AUM_31Jan.csv')
    aum_df = pd.read_csv(aum_path)
    name_to_isin = {}        # exact FUND_NAME -> ISIN
    norm_name_to_isin = {}   # normalized name -> ISIN (strips 'Fund', lowercases)
    for _, row in aum_df.iterrows():
        isin = row['ISIN']
        if pd.isna(isin):
            continue
        lookup.setdefault(isin, {})
        aum_val = row['AUM']
        lookup[isin]['aum_cr'] = aum_val / 10_000_000 if pd.notna(aum_val) else None
        fn = row.get('FUND_NAME')
        if pd.notna(fn) and fn:
            fn_str = str(fn).strip()
            name_to_isin[fn_str] = isin
            norm_name_to_isin[_norm_name(fn_str)] = isin

    # Store fallback maps inside lookup under reserved keys
    lookup['_name_to_isin'] = name_to_isin
    lookup['_norm_name_to_isin'] = norm_name_to_isin
    lookup['_isin_to_name'] = {isin: fn for fn, isin in name_to_isin.items()}

    # --- Power Rank ---
    pr_path = os.path.join(project_dir, 'Powerranking.csv')
    pr_df = pd.read_csv(pr_path)
    for _, row in pr_df.iterrows():
        isin = row['ISIN']
        if pd.isna(isin):
            continue
        lookup.setdefault(isin, {})
        pr = row['POWERRANK']
        lookup[isin]['powerrank'] = int(pr) if pd.notna(pr) else None

    # --- Upside / Downside ---
    ud_path = os.path.join(project_dir, 'upside_downside_mar.xlsx')
    wb = openpyxl.load_workbook(ud_path, data_only=True)
    ws = wb.active
    for r in range(2, ws.max_row + 1):
        isin = ws.cell(r, 2).value
        if not isin:
            continue
        lookup.setdefault(isin, {})
        down = ws.cell(r, 4).value
        up = ws.cell(r, 5).value
        lookup[isin]['downside'] = float(down) if (down and down != '--') else None
        lookup[isin]['upside'] = float(up) if (up and up != '--') else None

    # --- 5Y Rolling Returns ---
    rr_path = os.path.join(project_dir, 'Rolling_Returns_Mar.csv')
    rr_df = pd.read_csv(rr_path)
    df5 = rr_df[rr_df['ROLLING_PERIOD'] == 60]
    rr_scheme = {}    # ISIN -> 5Y return
    rr_category = {}  # UPDATED_SUBCATEGORY name -> 5Y category avg
    for _, row in df5.iterrows():
        eid = str(row['ENTITYID'])
        val = float(row['RETURN_VALUE'])
        if eid.startswith('INF'):
            rr_scheme[eid] = val
        else:
            rr_category[eid] = val

    # Merge scheme 5Y into main lookup
    for isin, ret in rr_scheme.items():
        lookup.setdefault(isin, {})
        lookup[isin]['five_year'] = ret

    print(f"  Reference data loaded: {len(lookup)} ISINs, {len(rr_category)} categories with 5Y avg")
    return lookup, rr_category


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------

def fmt_aum(aum_cr):
    """Format AUM in Crores with Indian comma notation. e.g. 76060.9 -> '76,061 Cr'"""
    if aum_cr is None:
        return '-'
    n = round(aum_cr)
    s = str(n)
    if len(s) <= 3:
        return f"{s} Cr"
    result = s[-3:]
    s = s[:-3]
    while s:
        result = s[-2:] + ',' + result
        s = s[:-2]
    return result.lstrip(',') + ' Cr'


def fmt_upside_downside(upside, downside):
    """Format as 'Upside | Downside' rounded to nearest integer. e.g. '93 | 89'"""
    if upside is None and downside is None:
        return '-'
    up_str = str(round(upside)) if upside is not None else '-'
    dn_str = str(round(downside)) if downside is not None else '-'
    return f"{up_str} | {dn_str}"


def fmt_powerrank(powerrank):
    """Format power rank as integer string."""
    if powerrank is None:
        return '-'
    return str(int(powerrank))
