"""
excel_reader.py — reads the PF_Curation sheet from the client Excel workbook.
Detects all four sections dynamically by header content.
"""
import openpyxl


def get_curation_sheet(wb):
    for name in wb.sheetnames:
        if name.startswith("PF_Curation"):
            return wb[name]
    raise ValueError("No PF_Curation* sheet found in workbook")


def col_map(header_row):
    """Return {column_name: 0-based_index} for non-None header values."""
    return {v: i for i, v in enumerate(header_row) if v is not None}


def detect_sections(ws):
    """
    Scan for section header rows by content.
    Returns dict with 's1','s2','s3','s4' as 1-based row numbers.

    How each section is identified:
      s1 — "Row Labels | Total Selected Value"  → Portfolio summary (AUM overview)
      s2 — FUND_NAME header row that contains FOLIO_NUMBER or TOTAL_UNITS
            → Current holdings / action master (per-folio sell/retain instructions)
      s3 — "Row Labels | Sum of TOTAL_VALUE" or "Sum of Current Value Amount"
            → Risk-group pivot summary
      s4 — FUND_NAME header row that contains "Allocation M1" or "Buy Value Amount"
            → Ideal portfolio (target allocation, buy/sell plan, milestones)
    If the smarter detection cannot tell s2 from s4, it falls back to:
      s2 = first FUND_NAME row, s4 = last FUND_NAME row.
    """
    sections = {}
    fund_name_rows = []   # (row_idx, header_tuple)
    _S3_B_VALUES = {"Sum of TOTAL_VALUE", "Sum of Current Value Amount"}

    all_rows = list(ws.iter_rows(values_only=True))
    for row_idx, row in enumerate(all_rows, start=1):
        a = row[0] if row else None
        b = row[1] if len(row) > 1 else None
        if a == "Row Labels" and b == "Total Selected Value":
            sections['s1'] = row_idx
        elif a == "Row Labels" and b in _S3_B_VALUES:
            sections['s3'] = row_idx
        elif a == "FUND_NAME":
            fund_name_rows.append((row_idx, row))

    # Identify s2 (action master) vs s4 (ideal portfolio) from FUND_NAME rows
    # s4 has columns like "Allocation M1", "Buy Value Amount", "SIP Allocation Amount"
    # s2 has columns like "FOLIO_NUMBER", "TOTAL_UNITS", "Sell Plan"
    _S4_MARKERS = {"Allocation M1", "Buy Value Amount", "SIP Allocation Amount"}
    _S2_MARKERS = {"FOLIO_NUMBER", "TOTAL_UNITS", "TOTAL_VALUE"}

    s2_candidates = []
    s4_candidates = []
    for row_idx, hdr in fund_name_rows:
        cols = set(c for c in hdr if c is not None)
        if cols & _S4_MARKERS:
            s4_candidates.append(row_idx)
        elif cols & _S2_MARKERS:
            s2_candidates.append(row_idx)

    if s4_candidates:
        sections['s4'] = s4_candidates[-1]   # last match is safest
    if s2_candidates:
        sections['s2'] = s2_candidates[0]    # first match

    # Fallback: first / last FUND_NAME row
    if 's2' not in sections and len(fund_name_rows) >= 1:
        sections['s2'] = fund_name_rows[0][0]
    if 's4' not in sections and len(fund_name_rows) >= 2:
        sections['s4'] = fund_name_rows[-1][0]

    return sections


def _read_section(ws, header_row_1based, col_names=None):
    """
    Read rows from header_row onwards until a blank col-A row or end of sheet.
    Returns list of dicts keyed by header name.
    Last non-blank row before the next blank section is tagged __grand_total__
    if its col-A contains 'Grand Total' or 'grand total'.
    """
    all_rows = list(ws.iter_rows(values_only=True))
    header = all_rows[header_row_1based - 1]
    cmap = col_map(header)
    if col_names:
        cmap = {k: v for k, v in cmap.items() if k in col_names}

    data = []
    for row in all_rows[header_row_1based:]:       # rows after header
        a = row[0] if row else None
        if a is None:
            break                                   # blank col-A = end of section
        record = {}
        for name, idx in cmap.items():
            record[name] = row[idx] if idx < len(row) else None
        label = str(a).strip().lower()
        if 'grand total' in label:
            record['__grand_total__'] = True
        data.append(record)
    return data


def _read_isin_column(ws, count):
    """
    Scan the sheet for a standalone 'ISIN' header cell (not in a multi-column header row),
    then read `count` ISIN values from the rows immediately below it in the same column.
    Returns a list of `count` strings (or None for missing values).
    The ISIN header row can vary per client — never hardcoded.
    """
    isin_header_row = None
    isin_col_0based = None

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for col_0, val in enumerate(row):
            if val == 'ISIN':
                isin_header_row = row_idx
                isin_col_0based = col_0
                break
        if isin_header_row is not None:
            break

    if isin_header_row is None:
        print("  WARNING: no standalone ISIN column found in sheet")
        return [None] * count

    all_rows = list(ws.iter_rows(values_only=True))
    isins = []
    for row in all_rows[isin_header_row: isin_header_row + count]:
        val = row[isin_col_0based] if isin_col_0based < len(row) else None
        isins.append(str(val).strip() if val else None)

    # Pad if fewer rows than expected
    while len(isins) < count:
        isins.append(None)

    print(f"  ISIN column found at row {isin_header_row}, col {isin_col_0based + 1}: {sum(1 for x in isins if x)} ISINs read")
    return isins


def _get_masterplan_sheet(wb):
    """Return the PF_MasterPlan_* sheet if present, else None."""
    for name in wb.sheetnames:
        if name.startswith('PF_MasterPlan'):
            return wb[name]
    return None


def _read_masterplan(ws):
    """
    Read the PF_MasterPlan_* sheet and return an excel_data dict compatible
    with what read_excel() produces for the old PF_Curation format.

    Sheet layout:
      Row 1 — title "Master Transition Plan"
      Row 2 — "PFV:" | <value>
      Row 3 — blank
      Row 4 — column headers (FUND_NAME, FOLIO_NUMBER, ISIN, …)
      Row 5+ — data rows (last row is Grand Total)

    Returns dict with section1/section2/section3/section4.
    ISIN is already a column in this sheet — no separate lookup needed.
    section2 and section3 are returned empty (not needed for this format).
    """
    all_rows = list(ws.iter_rows(values_only=True))

    # PFV from row 2 col B
    pfv = all_rows[1][1] if len(all_rows) > 1 and len(all_rows[1]) > 1 else 0
    pfv = pfv or 0

    # Headers at row 4 (index 3)
    header = all_rows[3]
    cmap = col_map(header)

    # Data rows from row 5 onwards; stop at first fully-blank row (col A None)
    section4 = []
    for row in all_rows[4:]:
        a = row[0] if row else None
        if a is None:
            break
        record = {}
        for name, idx in cmap.items():
            record[name] = row[idx] if idx < len(row) else None
        label = str(a).strip().lower()
        if 'grand total' in label:
            record['__grand_total__'] = True
        section4.append(record)

    # section1 — only needs Grand Total with Total Selected Value for PFV
    section1 = [{'Total Selected Value': pfv, '__grand_total__': True}]

    non_gt = [r for r in section4 if not r.get('__grand_total__')]
    print(f"  MasterPlan format detected: {len(non_gt)} data rows, PFV={pfv:,.0f}")
    return {
        'section1': section1,
        'section2': [],
        'section3': [],
        'section4': section4,
    }


def read_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # New format: PF_MasterPlan_* sheet takes priority
    mp_ws = _get_masterplan_sheet(wb)
    if mp_ws is not None:
        return _read_masterplan(mp_ws)

    # Old format: PF_Curation_* sheet
    ws = get_curation_sheet(wb)
    sections = detect_sections(ws)

    required = ['s1', 's2', 's3', 's4']
    missing = [k for k in required if k not in sections]
    if missing:
        raise ValueError(f"Could not detect sections: {missing}")

    result = {
        'section1': _read_section(ws, sections['s1']),
        'section2': _read_section(ws, sections['s2']),
        'section3': _read_section(ws, sections['s3']),
        'section4': _read_section(ws, sections['s4']),
    }

    # Attach ISINs to section4 rows — ISIN lives in a separate standalone column
    # positioned at a variable row; read positionally (same order as section4 rows)
    isins = _read_isin_column(ws, len(result['section4']))
    for row, isin in zip(result['section4'], isins):
        row['ISIN'] = isin

    print(f"  Sections found: {list(result.keys())}")
    for k, v in result.items():
        non_gt = [r for r in v if not r.get('__grand_total__')]
        print(f"    {k}: {len(non_gt)} data rows")
    return result
